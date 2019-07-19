from openpyxl import load_workbook
import csv

# funcao recursiva para preencher as 20 colunas:
# A2, B2, C2...
# A3, B3, C3...
def getInfo(linha, coluna, letra):
	letra2 = chr(letra + 1) # 64 + 1 == 65 --> 'A'
	cell = letra2 + str(linha) # celula a preencher é mescla da letra com numero da coluna
	print(cell) #imprimo todas as colunas de cada linha, separadas por linha tracejada
	ws[cell] = row[1][coluna] #jogo na celula o conteudo correspondente
	if coluna == 20: return # se for a ultima coluna retorna
	getInfo(linha, coluna + 1, letra + 1) # caso contrario, chama funcao novamente para preencher proxima coluna da mesma linha


wb = load_workbook('template.xlsx') #carrego o arquivo excel
ws = wb.active #escolho a planilha para trabalhar utilizando a aba ativa atual

#abrindo arq com with faz o fechamento automatico
with open('dados.csv', mode='r') as csv_file:
	csv_reader = csv.reader(csv_file, delimiter=';') # podemos definir o delimitador
	for row in enumerate(csv_reader): # usei enumerate para gerar indice para cada linha
		if row[0] == 0: continue # pra poder pular a primeira linha pois é definicao de colunas
		linha = int(row[0]) + 1 # entao pegamos a partir da segunda linha
		letra = 64 #'A' - 1 # utilizei a letra 'A' - 1 em numerico para poder incrementar
		getInfo(linha, 0, letra) # utilizo a letra como numerico
		print('-' * 20)

wb.save("sample.xlsx") #salvo planilha resultante
